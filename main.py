from openpyxl import load_workbook as lw
import itertools
import os


def main():
    global subjects_and_columns
    filename = get_filename()
    file = lw(filename, data_only=True)  # data_only - флаг для того, чтобы считывались значения формул
    page = file.active

    # end_row = page.max_row + 1
    end_row = 60
    end_column = page.max_column + 1

    subjects_and_columns = get_subjects(page, end_column)
    outcome = []
    scores = []

    intersections_limit = 22

    # Делим на 2 группы

    for L in range(1, len(subjects_and_columns.keys())):
        for subset in itertools.combinations(subjects_and_columns.keys(), L):
            list1 = list(subset)
            list2 = create_list2(subjects_and_columns.keys(), list1)
            score = find_score(list1, page, end_row) + find_score(list2, page, end_row)
            scores.append(score)
            if score < intersections_limit:
                list1.sort()
                list2.sort()
                for entry in outcome:
                    if list2 == entry[0]:
                        break
                else:
                    outcome.append([list1, list2, score])

    # Делим на 3 группы

    # for L in range(0, len(subjects_and_columns.keys())):
    #     for subset in itertools.combinations(subjects_and_columns.keys(), L):
    #         list1 = list(subset)
    #         list2 = create_list2(subjects_and_columns.keys(), list1)
    #         for N in range(0, len(list2)):
    #             for subset2 in itertools.combinations(list2, N):
    #                 list3 = list(subset2)
    #                 list4 = create_list2(list2, list3)
    #                 score = (find_score(list1, page, end_row) +
    #                          find_score(list3, page, end_row) +
    #                          find_score(list4, page, end_row))
    #                 scores.append(score)
    #                 if score < intersections_limit:
    #                     list1.sort()
    #                     list3.sort()
    #                     list4.sort()
    #                     for entry in outcome:
    #                         if list3 in entry and list4 in entry:
    #                             break
    #                     else:
    #                         outcome.append([list1, list3, list4, score])

    outcome.sort()
    for entry in outcome:
        print(entry)


def get_filename():
    """
    Показывает файлы в папке data, один из которых можно выбрать для обработки
    :return: str. Имя файла для дальнейшей работы
    """
    folder = 'data'
    files = os.listdir(folder)
    print(f'В папке {folder} найдены следующие файлы:')
    for i in range(len(files)):
        print(f'{i} - {files[i]}')
    choice = files[int(input('Введите номер файла, который нужно обработать: '))]
    return os.path.join(folder, choice)


def get_subjects(page, end):
    """
    Получает список профильных предметов
    Первые два столбца на странице - Фамилия и Имя, предмет начинается с 3 столбца
    :param page: объект страницы openpyxl
    :param end: последняя колонка, до которой ищутся предметы
    :return: dict. Ключ - название предмета, значение - номер колонки для этого предмета
    """
    subjects = {}
    for i in range(3, end):  # первый предмет находится в 3 столбце
        subj = page.cell(row=1, column=i).value
        subj = subj.strip()  # убираем лишние пробелы и квадратные скобки
        subjects[subj] = i
    # subjects.pop('Химия')
    subjects.pop('Немецкий')
    subjects.pop('История')
    subjects.pop('Математика углубленная')
    # subjects.pop('География')
    return subjects


def create_list2(main_list, list1):
    list2 = [i for i in main_list]
    for i in list1:
        list2.remove(i)
    return list2


def find_score(subjects, page, end):
    intersections = 0
    for i in range(1, end):
        score = 0
        for j in subjects:
            cell = page.cell(row=i, column=subjects_and_columns[j]).value
            if type(cell) is int:
                score += cell
        if score > 1:
            intersections += 1
    return intersections


if __name__ == '__main__':
    subjects_and_columns = None
    main()
