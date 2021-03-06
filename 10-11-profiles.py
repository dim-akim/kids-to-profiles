from openpyxl import load_workbook as lw
import itertools


filename = 'profiles-to-plans-21-22.xlsx'
columns = {
    'Физика': 2,
    'Химия': 3,
    'Биология': 4,
    'Обществознание': 5,
    'Литература': 6,
    'География': 7,
    'Информатика': 8,
    'Английский': 9
}
variant = [
    'Физика',
    'Химия',
    'Биология',
    'Обществознание',
    'Литература',
    'География',
    'Информатика',
    'Английский'
]


def main():
    file = lw(filename)
    page = file.active
    end = page.max_row + 1
    outcome = []
    scores = []

    for L in range(0, len(variant)):
        for subset in itertools.combinations(variant, L):
            list1 = list(subset)
            list2 = create_list2(list1)
            score = find_score(list1, page, end) + find_score(list2, page, end)
            scores.append(score)
            if score < 16:
                list1.sort()
                list2.sort()
                for entry in outcome:
                    if list2 == entry[0]:
                        break
                else:
                    outcome.append([list1, list2, score])

    for entry in outcome:
        print(entry)


def create_list2(list1):
    list2 = [i for i in variant]
    for i in list1:
        list2.remove(i)
    return list2


def find_score(subjects, page, end):
    intersections = 0
    for i in range(1, end):
        score = 0
        for j in subjects:
            cell = page.cell(row=i, column=columns[j]).value
            if type(cell) is int:
                score += cell
        if score > 1:
            intersections += 1
    return intersections


if __name__ == '__main__':
    main()
